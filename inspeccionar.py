"""
Script de inspeccion de PROYECTO.xlsx.

Ejecuta:  python inspeccionar.py

Analiza la hoja de seguimiento e imprime todo lo necesario para configurar
el MVP sin danar el archivo: hojas, encabezados, dimensiones, celdas
combinadas, formulas, formatos de fecha, tablas, autofiltros y paneles.
NO modifica el archivo (solo lectura).
"""
from pathlib import Path
import config
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

MAX_FILAS_MUESTRA = 12
MAX_COLS_MUESTRA = 12


def corto(v, n=28):
    if v is None:
        return ""
    s = str(v).replace("\n", " ")
    return s if len(s) <= n else s[: n - 1] + "…"


def inspeccionar():
    ruta = config.ARCHIVO_MAESTRO
    if not ruta.exists():
        print(f"[!] No encuentro el archivo: {ruta}")
        print("    Coloca PROYECTO.xlsx en esa ruta y vuelve a ejecutar.")
        return

    # data_only=False -> conserva y muestra las formulas tal cual
    wb = load_workbook(ruta, data_only=False)

    print("=" * 70)
    print(f"ARCHIVO: {ruta}")
    print(f"HOJAS: {wb.sheetnames}")
    print("=" * 70)

    nombre = config.HOJA_SEGUIMIENTO
    if nombre not in wb.sheetnames:
        print(f"[!] No existe una hoja llamada '{nombre}'.")
        print(f"    Ajusta HOJA_SEGUIMIENTO en config.py. Hojas: {wb.sheetnames}")
        # intenta adivinar
        for h in wb.sheetnames:
            if "segui" in h.lower():
                print(f"    Sugerencia: parece ser '{h}'")
        return

    ws = wb[nombre]
    print(f"\nHOJA ANALIZADA: '{nombre}'")
    print(f"  Dimensiones (max):  {ws.max_row} filas x {ws.max_column} columnas")
    print(f"  Rango declarado:    {ws.dimensions}")
    print(f"  Paneles congelados: {ws.freeze_panes}")
    print(f"  Autofiltro:         {ws.auto_filter.ref}")

    # Tablas (ListObjects) - importantes al agregar filas
    if ws.tables:
        print("\n  TABLAS (ListObjects) detectadas:")
        for tname, t in ws.tables.items():
            print(f"    - '{tname}'  rango={t.ref}")
        print("    OJO: si los datos estan dentro de una Tabla, al agregar filas")
        print("         nuevas hay que EXTENDER el rango de la tabla.")
    else:
        print("\n  No hay Tablas (ListObjects). Datos en rango normal.")

    # Celdas combinadas
    if ws.merged_cells.ranges:
        print(f"\n  CELDAS COMBINADAS ({len(ws.merged_cells.ranges)}):")
        for rng in list(ws.merged_cells.ranges)[:15]:
            print(f"    - {rng}")

    # Muestra de las primeras filas
    print(f"\n  MUESTRA (primeras {MAX_FILAS_MUESTRA} filas x {MAX_COLS_MUESTRA} cols):")
    ncols = min(ws.max_column, MAX_COLS_MUESTRA)
    header = "   fila | " + " | ".join(f"{get_column_letter(c):^12}" for c in range(1, ncols + 1))
    print("   " + header)
    for r in range(1, min(ws.max_row, MAX_FILAS_MUESTRA) + 1):
        celdas = []
        for c in range(1, ncols + 1):
            celdas.append(f"{corto(ws.cell(row=r, column=c).value, 12):^12}")
        print(f"   {r:>5} | " + " | ".join(celdas))

    # Formulas presentes en la hoja
    print("\n  FORMULAS detectadas (muestra):")
    formulas = []
    for fila in ws.iter_rows():
        for celda in fila:
            if isinstance(celda.value, str) and celda.value.startswith("="):
                formulas.append((celda.coordinate, celda.value))
    if formulas:
        for coord, f in formulas[:15]:
            print(f"    - {coord}: {corto(f, 45)}")
        cols_con_formula = sorted({c[0].rstrip("0123456789") for c in formulas})
        print(f"    Columnas con formula: {cols_con_formula}")
        print("    OJO: NO escribir en esas columnas para no romper calculos.")
    else:
        print("    (ninguna)")

    # Formatos de numero por columna (fila de datos)
    print("\n  FORMATOS de las columnas objetivo (usando primera fila de datos):")
    fila_datos = config.PRIMERA_FILA_DATOS
    for campo, letra in config.COLUMNAS.items():
        celda = ws[f"{letra}{fila_datos}"]
        print(f"    {campo:<15} col {letra}: valor={corto(celda.value,18)!r:<20} "
              f"formato={celda.number_format!r}")

    # Encabezados detectados
    print(f"\n  ENCABEZADOS (fila {config.FILA_ENCABEZADOS}):")
    for c in range(1, min(ws.max_column, 15) + 1):
        val = ws.cell(row=config.FILA_ENCABEZADOS, column=c).value
        if val is not None:
            print(f"    col {get_column_letter(c)} = {val!r}")

    print("\n" + "=" * 70)
    print("Revisa que el mapeo en config.py COLUMNAS coincida con lo de arriba.")
    print("=" * 70)


if __name__ == "__main__":
    inspeccionar()
