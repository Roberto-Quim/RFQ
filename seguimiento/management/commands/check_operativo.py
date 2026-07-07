"""
Chequeo operativo para uso interno (Fase 4).

    python manage.py check_operativo

Valida que el sistema este listo para uso interno controlado y termina con
codigo != 0 si hay problemas (util para scripts de arranque). NUNCA imprime el
valor de SECRET_KEY.
"""
from pathlib import Path

from django.conf import settings
from django.contrib.auth.models import Permission
from django.core.management.base import BaseCommand, CommandError
from django.db import connection
from django.db.migrations.executor import MigrationExecutor

import config

# Claves de ejemplo que NO deben usarse en modo interno.
SECRETS_INSEGUROS = {
    "",
    "dev-inseguro-solo-local-cambiar-en-produccion",
    "clave-ficticia",
    "pon-aqui-una-clave-larga-y-secreta",
}


class Command(BaseCommand):
    help = "Chequeo operativo del sistema RFQ para uso interno."

    def handle(self, *args, **options):
        ok, problemas = [], []

        def marca(cond, texto_ok, texto_error):
            (ok if cond else problemas).append(texto_ok if cond else texto_error)

        # DEBUG
        marca(not settings.DEBUG, "DEBUG=False",
              "DEBUG=True: para uso interno usa DJANGO_DEBUG=0")

        # SECRET_KEY (sin exponer el valor)
        marca(settings.SECRET_KEY not in SECRETS_INSEGUROS,
              "SECRET_KEY personalizada",
              "SECRET_KEY es de ejemplo/insegura: define DJANGO_SECRET_KEY")

        # PROYECTO.xlsx y hoja
        if config.ARCHIVO_MAESTRO.exists():
            ok.append(f"PROYECTO.xlsx existe ({config.ARCHIVO_MAESTRO})")
            try:
                from openpyxl import load_workbook
                wb = load_workbook(config.ARCHIVO_MAESTRO, read_only=True)
                hojas = wb.sheetnames
                wb.close()
                marca(config.HOJA_SEGUIMIENTO in hojas,
                      f"Hoja '{config.HOJA_SEGUIMIENTO}' existe",
                      f"Falta la hoja '{config.HOJA_SEGUIMIENTO}' (hojas: {hojas})")
            except Exception as e:  # noqa: BLE001
                problemas.append(f"No se pudo leer el maestro: {e}")
        else:
            problemas.append(f"No existe PROYECTO.xlsx en {config.ARCHIVO_MAESTRO}")

        # Carpetas (se crean si faltan)
        carpetas = {
            "media": settings.MEDIA_ROOT,
            "backups": config.CARPETA_BACKUPS,
            "reportes": config.CARPETA_REPORTES,
            "logs": settings.LOGS_DIR,
        }
        for etiqueta, ruta in carpetas.items():
            try:
                Path(ruta).mkdir(parents=True, exist_ok=True)
                ok.append(f"carpeta {etiqueta} lista ({ruta})")
            except Exception as e:  # noqa: BLE001
                problemas.append(f"no se pudo crear carpeta {etiqueta}: {e}")

        # Base de datos
        try:
            connection.ensure_connection()
            ok.append("base de datos responde")
        except Exception as e:  # noqa: BLE001
            problemas.append(f"base de datos no responde: {e}")

        # Migraciones aplicadas
        try:
            executor = MigrationExecutor(connection)
            pendientes = executor.migration_plan(executor.loader.graph.leaf_nodes())
            marca(not pendientes, "migraciones al dia",
                  f"hay {len(pendientes)} migracion(es) pendiente(s): ejecuta migrate")
        except Exception as e:  # noqa: BLE001
            problemas.append(f"no se pudo verificar migraciones: {e}")

        # Permiso de confirmar
        try:
            existe = Permission.objects.filter(codename="puede_confirmar").exists()
        except Exception:  # noqa: BLE001
            existe = False
        marca(existe, "permiso 'puede_confirmar' existe",
              "falta el permiso 'puede_confirmar': ejecuta migrate")

        # Salida
        for o in ok:
            self.stdout.write(self.style.SUCCESS(f"[OK] {o}"))
        for p in problemas:
            self.stdout.write(self.style.ERROR(f"[X]  {p}"))

        if problemas:
            raise CommandError(f"{len(problemas)} problema(s) detectado(s). "
                               "Revisa la lista de arriba.")
        self.stdout.write(self.style.SUCCESS("\nTodo listo para uso interno."))
