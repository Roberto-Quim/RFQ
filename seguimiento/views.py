"""
Vistas del flujo web seguro.

  subir_rfq  -> sube .txt y muestra PREVIEW editable (NO escribe al Excel).
  confirmar  -> escribe al Excel (requiere permiso) y guarda auditoria.
  historial  -> lista los RFQ procesados.
  estado     -> panel de estado del sistema (sin exponer secretos).

Modo seguro:
  - La escritura solo ocurre en 'confirmar', tras revisar/editar el preview.
  - 'confirmar' exige el permiso 'seguimiento.puede_confirmar'. Sin el, el
    usuario puede subir y previsualizar, pero no escribir al Excel.
"""
import csv
import logging

from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import redirect, render

import config
from .forms import SubirRFQForm, VistaPreviaForm
from .models import RFQProcesado
from .services import MotorError, escribir_confirmado, extraer_preview, guardar_subida
from src.modelo import RFQData

log = logging.getLogger("seguimiento")

PERM_CONFIRMAR = "seguimiento.puede_confirmar"


def _iso(fecha):
    return fecha.isoformat() if fecha else ""


@login_required
def subir_rfq(request):
    """Paso 1-3: sube el .txt, extrae datos y muestra la vista previa editable."""
    if request.method == "POST":
        form = SubirRFQForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = form.cleaned_data["archivo"]
            ruta = guardar_subida(archivo, settings.MEDIA_ROOT)
            try:
                dato = extraer_preview(ruta)
            except MotorError as e:
                messages.error(request, str(e))
                return render(request, "seguimiento/subir_rfq.html", {"form": form})

            preview = VistaPreviaForm(initial={
                "rfq": dato.rfq,
                "descripcion": dato.descripcion or "",
                "fecha_arranque": dato.fecha_arranque,
                "solicitante": dato.solicitante or "",
                "planta": dato.planta or "",
                "archivo_nombre": archivo.name,
                # originales para auditar ediciones:
                "orig_rfq": dato.rfq,
                "orig_descripcion": dato.descripcion or "",
                "orig_fecha_arranque": _iso(dato.fecha_arranque),
                "orig_solicitante": dato.solicitante or "",
                "orig_planta": dato.planta or "",
            })
            return render(request, "seguimiento/vista_previa.html", {
                "form": preview,
                "faltantes": dato.faltantes,
                "archivo_nombre": archivo.name,
                "puede_confirmar": request.user.has_perm(PERM_CONFIRMAR),
            })
    else:
        form = SubirRFQForm()
    return render(request, "seguimiento/subir_rfq.html", {"form": form})


@login_required
def confirmar(request):
    """Paso 4-6: escribe al Excel SOLO tras confirmar y con permiso; audita."""
    if request.method != "POST":
        return redirect("subir_rfq")

    # Control de permiso: sin el, puede ver preview pero no escribir.
    if not request.user.has_perm(PERM_CONFIRMAR):
        log.warning("Usuario '%s' intento confirmar sin permiso", request.user)
        return render(request, "seguimiento/resultado.html", {
            "ok": False,
            "mensaje": "No tienes permiso para confirmar la escritura al Excel. "
                       "Solicita el permiso 'Confirmadores RFQ'.",
        }, status=403)

    form = VistaPreviaForm(request.POST)
    if not form.is_valid():
        return render(request, "seguimiento/vista_previa.html", {
            "form": form,
            "faltantes": [],
            "archivo_nombre": request.POST.get("archivo_nombre", ""),
            "puede_confirmar": True,
        })

    cd = form.cleaned_data
    editados = form.campos_editados()
    dato = RFQData(
        rfq=cd["rfq"].strip(),
        descripcion=(cd["descripcion"] or None),
        fecha_arranque=(cd["fecha_arranque"] or None),
        solicitante=(cd["solicitante"] or None),
        planta=(cd["planta"] or None),
        origen_archivo=cd.get("archivo_nombre", ""),
    )
    dato.registrar_faltantes()

    try:
        resumen = escribir_confirmado(dato)
    except MotorError as e:
        RFQProcesado.objects.create(
            rfq=dato.rfq, descripcion=dato.descripcion or "",
            fecha_arranque=dato.fecha_arranque, solicitante=dato.solicitante or "",
            planta=dato.planta or "", archivo_nombre=dato.origen_archivo,
            estado=RFQProcesado.ESTADO_ERROR, accion="",
            campos_faltantes=", ".join(dato.faltantes),
            campos_editados=", ".join(editados),
            usuario=request.user, mensaje=str(e),
        )
        return render(request, "seguimiento/resultado.html", {
            "ok": False, "mensaje": str(e), "dato": dato,
        })

    RFQProcesado.objects.create(
        rfq=dato.rfq, descripcion=dato.descripcion or "",
        fecha_arranque=dato.fecha_arranque, solicitante=dato.solicitante or "",
        planta=dato.planta or "", archivo_nombre=dato.origen_archivo,
        estado=RFQProcesado.ESTADO_OK, accion=resumen["accion"],
        campos_faltantes=", ".join(resumen["faltantes"]),
        campos_editados=", ".join(editados),
        usuario=request.user, mensaje=f"Backup: {resumen['backup']}",
    )
    return render(request, "seguimiento/resultado.html", {
        "ok": True, "dato": dato, "resumen": resumen, "editados": editados,
    })


@login_required
def historial(request):
    """Lista los ultimos RFQ procesados."""
    registros = RFQProcesado.objects.select_related("usuario")[:100]
    return render(request, "seguimiento/historial.html", {"registros": registros})


# Columnas de la exportacion CSV (encabezado y orden).
CSV_COLUMNAS = [
    "fecha", "usuario", "rfq", "descripcion", "solicitante", "planta",
    "accion", "estado", "campos_faltantes", "campos_editados", "mensaje",
]


@login_required
def exportar_historial_csv(request):
    """Descarga el historial completo como CSV (requiere login)."""
    respuesta = HttpResponse(content_type="text/csv; charset=utf-8")
    respuesta["Content-Disposition"] = 'attachment; filename="historial_rfq.csv"'
    respuesta.write("﻿")  # BOM para que Excel abra los acentos bien

    escritor = csv.writer(respuesta)
    escritor.writerow(CSV_COLUMNAS)
    for r in RFQProcesado.objects.select_related("usuario").all():
        escritor.writerow([
            r.creado_en.strftime("%Y-%m-%d %H:%M"),
            r.usuario.get_username() if r.usuario else "",
            r.rfq, r.descripcion, r.solicitante, r.planta,
            r.accion, r.estado, r.campos_faltantes, r.campos_editados, r.mensaje,
        ])
    return respuesta


@login_required
def estado(request):
    """Panel de estado del sistema. NO expone secretos (SECRET_KEY, etc.)."""
    hoja_ok = False
    hojas = []
    if config.ARCHIVO_MAESTRO.exists():
        try:
            from openpyxl import load_workbook
            wb = load_workbook(config.ARCHIVO_MAESTRO, read_only=True)
            hojas = wb.sheetnames
            hoja_ok = config.HOJA_SEGUIMIENTO in hojas
            wb.close()
        except Exception as e:  # noqa: BLE001
            log.error("No se pudo inspeccionar el maestro: %s", e)

    # Chequeo simple de BD.
    try:
        RFQProcesado.objects.exists()
        db_ok = True
    except Exception:  # noqa: BLE001
        db_ok = False

    contexto = {
        "maestro_ruta": str(config.ARCHIVO_MAESTRO),
        "maestro_existe": config.ARCHIVO_MAESTRO.exists(),
        "hoja_nombre": config.HOJA_SEGUIMIENTO,
        "hoja_existe": hoja_ok,
        "hojas": hojas,
        "media_existe": settings.MEDIA_ROOT.exists(),
        "backups_existe": config.CARPETA_BACKUPS.exists(),
        "reportes_existe": config.CARPETA_REPORTES.exists(),
        "logs_existe": settings.LOGS_DIR.exists(),
        "db_ok": db_ok,
        "debug": settings.DEBUG,
        "allowed_hosts": settings.ALLOWED_HOSTS,
        "secret_key_definida": bool(settings.SECRET_KEY),  # solo bool, nunca el valor
    }
    return render(request, "seguimiento/estado.html", contexto)
