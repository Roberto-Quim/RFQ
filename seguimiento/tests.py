"""
Pruebas de integracion del flujo web (Fase 2 + Fase 3) con datos FICTICIOS.

Usan la BD de prueba de Django (no tocan db.sqlite3) y un PROYECTO.xlsx
TEMPORAL (no tocan el Excel real). Ejecuta:  python manage.py test seguimiento
"""
import shutil
import tempfile
from io import StringIO
from pathlib import Path

from django.conf import settings
from django.contrib.auth.models import Permission, User
from django.core.management import call_command
from django.core.management.base import CommandError
from django.test import Client, TestCase
from openpyxl import Workbook, load_workbook

import config
from seguimiento.locking import LockOcupado, lock_escritura_excel
from seguimiento.models import RFQProcesado


class FlujoWebTests(TestCase):
    def setUp(self):
        self.tmp = Path(tempfile.mkdtemp())
        self._orig_master = config.ARCHIVO_MAESTRO
        self._orig_backups = config.CARPETA_BACKUPS
        config.ARCHIVO_MAESTRO = self.tmp / "PROYECTO.xlsx"
        config.CARPETA_BACKUPS = self.tmp / "backups"
        self._crear_maestro_temporal()

        # Usuario CON permiso de confirmar.
        self.confirmador = User.objects.create_user("confirmador", password="x")
        perm = Permission.objects.get(codename="puede_confirmar")
        self.confirmador.user_permissions.add(perm)
        self.cli = Client()
        self.cli.force_login(self.confirmador)

        # Usuario SIN permiso.
        self.lector = User.objects.create_user("lector", password="x")
        self.cli_lector = Client()
        self.cli_lector.force_login(self.lector)

    def tearDown(self):
        config.ARCHIVO_MAESTRO = self._orig_master
        config.CARPETA_BACKUPS = self._orig_backups
        shutil.rmtree(self.tmp, ignore_errors=True)

    def _crear_maestro_temporal(self):
        wb = Workbook()
        ws = wb.active
        ws.title = config.HOJA_SEGUIMIENTO  # "SEGUIMIENTO " con espacio
        headers = {1: "RFQ", 2: "DESCRIPCION", 3: "FECHA DE ARRANQUE",
                   4: "SOLICITANTE", 6: "PLANTA", 15: "O", 16: "P"}
        for c, t in headers.items():
            ws.cell(row=2, column=c, value=t)
        ws.cell(row=3, column=1, value="100").number_format = "@"
        ws.cell(row=3, column=15, value='=IF(N3<=M3,"SI","NO")')
        ws.cell(row=3, column=16, value="=N3-C3")
        wb.save(config.ARCHIVO_MAESTRO)

    def _post_confirmar(self, cli, rfq="228", fecha=""):
        return cli.post("/confirmar/", {
            "rfq": rfq, "descripcion": "Rectificadora plana",
            "solicitante": "Jose Manuel Martinez", "planta": "Ramos",
            "fecha_arranque": fecha, "archivo_nombre": "pedido_228.txt",
            # originales extraidos (para auditar ediciones):
            "orig_rfq": rfq, "orig_descripcion": "Rectificadora plana",
            "orig_fecha_arranque": "", "orig_solicitante": "Jose Manuel Martinez",
            "orig_planta": "Ramos",
        })

    # --- Fase 2 (compatibilidad) ---
    def test_login_requerido(self):
        self.assertEqual(Client().get("/").status_code, 302)

    def test_confirmar_escribe_y_registra(self):
        # fecha editada a mano (orig venia vacia) -> debe quedar auditada.
        resp = self._post_confirmar(self.cli, fecha="2026-07-01")
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "agregado")

        ws = load_workbook(config.ARCHIVO_MAESTRO)[config.HOJA_SEGUIMIENTO]
        self.assertEqual(ws["A4"].value, "228")
        self.assertEqual(ws["A4"].number_format, "@")
        self.assertEqual(ws["O4"].value, '=IF(N4<=M4,"SI","NO")')
        self.assertEqual(ws["P4"].value, "=N4-C4")

        reg = RFQProcesado.objects.get(rfq="228")
        self.assertEqual(reg.estado, "ok")
        self.assertEqual(reg.accion, "agregado")
        self.assertEqual(reg.usuario, self.confirmador)          # auditoria de usuario
        self.assertIn("fecha_arranque", reg.campos_editados)     # auditoria de edicion

    def test_confirmar_sin_maestro_da_error_claro(self):
        config.ARCHIVO_MAESTRO.unlink()
        resp = self._post_confirmar(self.cli, rfq="999")
        self.assertContains(resp, "No existe")
        reg = RFQProcesado.objects.get(rfq="999")
        self.assertEqual(reg.estado, "error")
        self.assertEqual(reg.usuario, self.confirmador)

    # --- Fase 3: permisos ---
    def test_sin_permiso_no_confirma(self):
        resp = self._post_confirmar(self.cli_lector, rfq="555")
        self.assertEqual(resp.status_code, 403)
        # No debe haberse escrito ni registrado nada.
        self.assertFalse(RFQProcesado.objects.filter(rfq="555").exists())
        ws = load_workbook(config.ARCHIVO_MAESTRO)[config.HOJA_SEGUIMIENTO]
        self.assertIsNone(ws["A4"].value)

    def test_con_permiso_si_confirma(self):
        resp = self._post_confirmar(self.cli, rfq="556")
        self.assertEqual(resp.status_code, 200)
        self.assertTrue(RFQProcesado.objects.filter(rfq="556", estado="ok").exists())

    # --- Fase 3: pagina de estado ---
    def test_estado_requiere_login(self):
        self.assertEqual(Client().get("/estado/").status_code, 302)

    def test_estado_no_expone_secreto(self):
        from django.conf import settings
        resp = self.cli.get("/estado/")
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "Estado del sistema")
        self.assertNotContains(resp, settings.SECRET_KEY)  # nunca el valor real


class LockTests(TestCase):
    def test_lock_bloquea_si_ocupado(self):
        tmp = Path(tempfile.mkdtemp())
        try:
            ruta = tmp / ".lock"
            ruta.write_text("otro-proceso")  # simula lock de otro proceso, reciente
            with self.assertRaises(LockOcupado):
                with lock_escritura_excel(ruta, timeout=0.4, espera=0.1):
                    pass  # no deberia llegar aqui
        finally:
            shutil.rmtree(tmp, ignore_errors=True)

    def test_lock_se_adquiere_y_libera(self):
        tmp = Path(tempfile.mkdtemp())
        try:
            ruta = tmp / ".lock"
            with lock_escritura_excel(ruta, timeout=1):
                self.assertTrue(ruta.exists())   # lock tomado
            self.assertFalse(ruta.exists())      # liberado al salir
        finally:
            shutil.rmtree(tmp, ignore_errors=True)


class Fase4Tests(TestCase):
    """Pruebas de despliegue interno (Waitress, estaticos, chequeo operativo)."""

    def test_static_root_configurado(self):
        self.assertTrue(str(settings.STATIC_ROOT).endswith("staticfiles"))
        self.assertIn(
            "whitenoise",
            settings.STORAGES["staticfiles"]["BACKEND"].lower(),
        )

    def test_run_waitress_importa_app_sin_arrancar(self):
        import run_waitress
        app = run_waitress.get_application()
        self.assertTrue(callable(app))  # es la app WSGI; no arranco servidor

    def test_check_operativo_detecta_maestro_faltante(self):
        tmp = Path(tempfile.mkdtemp())
        orig = config.ARCHIVO_MAESTRO
        config.ARCHIVO_MAESTRO = tmp / "NO_EXISTE.xlsx"
        try:
            out = StringIO()
            with self.assertRaises(CommandError):
                call_command("check_operativo", stdout=out, stderr=StringIO())
            self.assertIn("No existe PROYECTO.xlsx", out.getvalue())
        finally:
            config.ARCHIVO_MAESTRO = orig
            shutil.rmtree(tmp, ignore_errors=True)

    def test_check_operativo_no_expone_secreto(self):
        out = StringIO()
        try:
            call_command("check_operativo", stdout=out, stderr=StringIO())
        except CommandError:
            pass
        self.assertNotIn(settings.SECRET_KEY, out.getvalue())


class Fase5Tests(TestCase):
    """Operacion persistente: rotacion de logs, --simple y scripts."""

    def test_logging_usa_rotating_file_handler(self):
        h = settings.LOGGING["handlers"]["archivo"]
        self.assertTrue(h["class"].endswith("RotatingFileHandler"))
        self.assertEqual(h["backupCount"], 5)
        self.assertEqual(h["maxBytes"], 5 * 1024 * 1024)

    def test_check_operativo_simple_no_expone_secreto(self):
        out = StringIO()
        try:
            call_command("check_operativo", "--simple", stdout=out, stderr=StringIO())
        except CommandError:
            pass
        salida = out.getvalue()
        self.assertIn("RFQ CHECK", salida)                 # salida compacta
        self.assertNotIn(settings.SECRET_KEY, salida)      # nunca el valor

    def test_scripts_operativos_existen(self):
        scripts = Path(settings.BASE_DIR) / "scripts"
        for nombre in ("iniciar_servidor.bat", "iniciar_servidor.ps1",
                       "check_operativo.bat", "migrar_y_collectstatic.bat",
                       "README_OPERACION_WINDOWS.md"):
            self.assertTrue((scripts / nombre).exists(), f"falta scripts/{nombre}")

    def test_run_waitress_sigue_importable(self):
        import run_waitress
        self.assertTrue(callable(run_waitress.get_application()))
