"""
Extractor de RFQ en Excel (.xlsx / .xlsm).

ESTADO: plantilla lista para adaptar. Como todavia NO conocemos el formato
real de los RFQ en Excel, aqui hay dos estrategias posibles. Deja la que
aplique cuando veamos archivos reales:

  A) RFQ con CELDAS ETIQUETADAS: en algun lado dice "RFQ:", "Planta:", etc.
     -> se busca la etiqueta y se toma la celda de al lado. (implementado abajo)

  B) RFQ en formato TABLA: una fila por RFQ con encabezados de columna.
     -> se leeria con pandas/openpyxl como tabla. (pendiente)

Ajusta ETIQUETAS con los textos reales que aparezcan en los archivos.
"""
from pathlib import Path

from openpyxl import load_workbook

from src.extractores.base import Extractor
from src.extractores.utilidades import parsear_fecha, limpiar_texto
from src.modelo import RFQData

# Sinonimos de etiqueta -> campo interno. AJUSTAR con textos reales.
ETIQUETAS = {
    "rfq": ("rfq", "no. rfq", "numero de rfq", "folio"),
    "descripcion": ("descripcion", "descripción", "proyecto", "parte"),
    "fecha_arranque": ("fecha de arranque", "arranque", "sop", "start of production"),
    "solicitante": ("solicitante", "solicita", "contacto", "requestor"),
    "planta": ("planta", "plant", "ubicacion", "ubicación"),
}


def _campo_de_etiqueta(texto: str):
    t = str(texto).strip().lower().rstrip(":")
    for campo, sinonimos in ETIQUETAS.items():
        if t in sinonimos:
            return campo
    return None


class ExcelRfqExtractor(Extractor):
    extensiones = (".xlsx", ".xlsm")

    def extraer(self, path: Path) -> list[RFQData]:
        wb = load_workbook(path, data_only=True)  # queremos valores, no formulas
        ws = wb.active
        campos = {"rfq": None, "descripcion": None, "fecha_arranque": None,
                  "solicitante": None, "planta": None}

        # Estrategia A: buscar etiquetas y tomar el valor de la celda contigua.
        for fila in ws.iter_rows():
            for celda in fila:
                if celda.value is None:
                    continue
                campo = _campo_de_etiqueta(celda.value)
                if campo and campos[campo] is None:
                    vecino = ws.cell(row=celda.row, column=celda.column + 1).value
                    campos[campo] = vecino

        rfq = limpiar_texto(campos["rfq"])
        if not rfq:
            # No se encontro un RFQ: se reporta arriba como archivo sin datos.
            return []

        dato = RFQData(
            rfq=rfq,
            descripcion=limpiar_texto(campos["descripcion"]),
            fecha_arranque=parsear_fecha(campos["fecha_arranque"]),
            solicitante=limpiar_texto(campos["solicitante"]),
            planta=limpiar_texto(campos["planta"]),
            origen_archivo=path.name,
        )
        dato.registrar_faltantes()
        return [dato]
