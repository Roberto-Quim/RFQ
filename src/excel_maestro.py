"""
Lectura y escritura de la hoja SEGUIMIENTO en PROYECTO.xlsx.

Principios:
  - Solo se tocan las celdas de las columnas mapeadas (A, B, C, D, F).
  - Nunca se reescribe la fila completa: las demas columnas quedan intactas.
  - RFQ SIEMPRE como texto (formato '@'), para conservar "131/4", "185.5", etc.
  - La fecha se escribe como fecha REAL de Excel con su number_format.
  - Antes de guardar se hace backup y se escribe de forma atomica (temp + replace)
    para no corromper el archivo si algo falla a mitad del guardado.
"""
import os
import shutil
from copy import copy
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.utils import column_index_from_string, get_column_letter

import config


# ----------------------------------------------------------------------
# Normalizacion de RFQ (para comparar y evitar duplicados de forma robusta)
# ----------------------------------------------------------------------
def normalizar_rfq(valor) -> str:
    """Convierte cualquier representacion de RFQ a texto canonico para comparar.

    Ejemplos:
        185      -> "185"
        185.0    -> "185"     (evita el .0 que mete Excel al leer numeros)
        185.5    -> "185.5"
        "131/4 " -> "131/4"
    """
    if valor is None:
        return ""
    if isinstance(valor, float) and valor.is_integer():
        valor = int(valor)
    return str(valor).strip()


# ----------------------------------------------------------------------
# Carga
# ----------------------------------------------------------------------
def cargar():
    """Abre el libro conservando formulas, estilos, filtros y tablas."""
    if not config.ARCHIVO_MAESTRO.exists():
        raise FileNotFoundError(f"No existe el maestro: {config.ARCHIVO_MAESTRO}")
    wb = load_workbook(config.ARCHIVO_MAESTRO)  # data_only=False por defecto
    if config.HOJA_SEGUIMIENTO not in wb.sheetnames:
        raise KeyError(
            f"No existe la hoja '{config.HOJA_SEGUIMIENTO}'. "
            f"Hojas: {wb.sheetnames}. Ajusta config.HOJA_SEGUIMIENTO."
        )
    return wb, wb[config.HOJA_SEGUIMIENTO]


def _col(campo: str) -> int:
    """Indice numerico de columna a partir del mapeo de config."""
    return column_index_from_string(config.COLUMNAS[campo])


def ultima_fila_datos(ws) -> int:
    """Ultima fila que realmente tiene un RFQ en la columna A.

    ws.max_row no sirve: suele venir inflado por formato/bordes en filas vacias.
    """
    col_rfq = _col("rfq")
    ultima = config.PRIMERA_FILA_DATOS - 1
    for r in range(config.PRIMERA_FILA_DATOS, ws.max_row + 1):
        if ws.cell(row=r, column=col_rfq).value not in (None, ""):
            ultima = r
    return ultima


# ----------------------------------------------------------------------
# BUSCAR: localizar la fila de un RFQ existente
# ----------------------------------------------------------------------
def buscar_fila_rfq(ws, rfq: str):
    """Devuelve el numero de fila donde esta el RFQ, o None si no existe."""
    objetivo = normalizar_rfq(rfq)
    col_rfq = _col("rfq")
    for r in range(config.PRIMERA_FILA_DATOS, ultima_fila_datos(ws) + 1):
        if normalizar_rfq(ws.cell(row=r, column=col_rfq).value) == objetivo:
            return r
    return None


# ----------------------------------------------------------------------
# ESCRIBIR: agregar o actualizar una fila (solo columnas mapeadas)
# ----------------------------------------------------------------------
def _escribir_celda(ws, fila: int, campo: str, valor):
    """Escribe UNA celda respetando tipo y formato segun el campo."""
    celda = ws.cell(row=fila, column=_col(campo))

    if campo == "rfq":
        # Forzar texto: primero el formato, luego el valor como str.
        celda.number_format = "@"
        celda.value = normalizar_rfq(valor)
        return

    if campo == "fecha_arranque":
        if isinstance(valor, (date, datetime)):
            celda.value = valor
            celda.number_format = config.FORMATO_FECHA
        elif valor in (None, ""):
            celda.value = config.VALOR_FALTANTE
        else:
            # vino como texto no parseable: se guarda tal cual, se avisa arriba
            celda.value = valor
        return

    # descripcion, solicitante, planta
    celda.value = valor if valor not in (None, "") else config.VALOR_FALTANTE


def _copiar_estilo_fila(ws, origen: int, destino: int):
    """Prepara una fila nueva heredando de la fila anterior (origen):

    1. Copia el estilo/formato (bordes, formato de numero) en TODAS las columnas.
    2. En columnas NO gestionadas que tengan formula (O, P, ...), replica la
       formula ajustando sus referencias a la fila nueva con el traductor de
       openpyxl (maneja bien referencias relativas/absolutas; no rompe numeros).
    3. NO copia el VALOR de las columnas gestionadas (A, B, C, D, F): esas las
       escribe _escribir_celda, para no arrastrar el dato del renglon anterior.
    """
    gestionadas = {_col(c) for c in config.COLUMNAS}
    for c in range(1, ws.max_column + 1):
        c_orig = ws.cell(row=origen, column=c)
        c_dest = ws.cell(row=destino, column=c)

        # 1) heredar estilo/formato en todas las columnas
        if c_orig.has_style:
            c_dest._style = copy(c_orig._style)

        # 2) las gestionadas se escriben aparte; no copiar su valor aqui
        if c in gestionadas:
            continue

        # 3) columnas no gestionadas: solo replicar formulas (ej. O y P),
        #    ajustadas a la fila destino. Los valores no-formula no se copian.
        if isinstance(c_orig.value, str) and c_orig.value.startswith("="):
            origen_coord = f"{get_column_letter(c)}{origen}"
            destino_coord = f"{get_column_letter(c)}{destino}"
            c_dest.value = Translator(c_orig.value, origin=origen_coord).translate_formula(destino_coord)

    # alto de fila
    if origen in ws.row_dimensions:
        ws.row_dimensions[destino].height = ws.row_dimensions[origen].height


def agregar_o_actualizar(ws, dato) -> str:
    """Punto de entrada por cada RFQData. Devuelve 'agregado' o 'actualizado'."""
    fila = buscar_fila_rfq(ws, dato.rfq)

    if fila is None:
        # --- AGREGAR al final ---
        fila_previa = ultima_fila_datos(ws)
        fila = fila_previa + 1
        if config.COPIAR_ESTILO_FILA_NUEVA and fila_previa >= config.PRIMERA_FILA_DATOS:
            _copiar_estilo_fila(ws, fila_previa, fila)
        resultado = "agregado"
    else:
        resultado = "actualizado"

    # Escribir SOLO las columnas mapeadas. Las demas no se tocan.
    _escribir_celda(ws, fila, "rfq", dato.rfq)
    _escribir_celda(ws, fila, "descripcion", dato.descripcion)
    _escribir_celda(ws, fila, "fecha_arranque", dato.fecha_arranque)
    _escribir_celda(ws, fila, "solicitante", dato.solicitante)
    _escribir_celda(ws, fila, "planta", dato.planta)

    return resultado


# ----------------------------------------------------------------------
# GUARDAR: backup + escritura atomica
# ----------------------------------------------------------------------
def backup() -> Path:
    """Copia el maestro actual a backups/ antes de modificarlo."""
    config.CARPETA_BACKUPS.mkdir(exist_ok=True)
    marca = datetime.now().strftime("%Y%m%d_%H%M%S")
    destino = config.CARPETA_BACKUPS / f"PROYECTO_{marca}.xlsx"
    shutil.copy2(config.ARCHIVO_MAESTRO, destino)
    return destino


def guardar(wb):
    """Guarda de forma atomica: escribe a un .tmp y reemplaza el original."""
    tmp = config.ARCHIVO_MAESTRO.with_suffix(".tmp.xlsx")
    wb.save(tmp)
    os.replace(tmp, config.ARCHIVO_MAESTRO)  # reemplazo atomico en el mismo disco
