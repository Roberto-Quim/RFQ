"""
Validacion minima del motor de Fase 1. NO usa datos reales ni PROYECTO.xlsx.

Ejecuta:
    python tests/test_motor.py

Construye una hoja SINTETICA que imita la estructura real:
  - nombre de hoja con espacio final ("SEGUIMIENTO "),
  - encabezados en fila 2, datos desde fila 3,
  - formulas en columnas O y P.
Y verifica las garantias clave:
  - RFQ desde PEDIDO #/Request # (extractor de correo),
  - RFQ guardado como TEXTO,
  - no duplica: actualiza si existe, agrega si no,
  - solo escribe A/B/C/D/F, no toca O/P al actualizar,
  - al agregar fila nueva copia y AJUSTA las formulas de O y P.
"""
import sys
import tempfile
from datetime import date
from pathlib import Path

# permitir "import config" y "from src..." al correr desde tests/
RAIZ = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(RAIZ))

import config  # noqa: E402
from openpyxl import Workbook  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402
from src import excel_maestro as maestro  # noqa: E402
from src.extractores.correo_formapprovals import FormApprovalsExtractor, parsear_texto  # noqa: E402
from src.modelo import RFQData  # noqa: E402

CORREO_FICTICIO = """PEDIDO #999 | DESTINATARIOS: 4 de 4 | JUL 07, 2026

RFQ - CapEx Interno
The request is now complete.

Nombre del Solicitante:
Persona Ficticia Uno
Selecciona Unidad de Negocio:
Planta Demo Norte
Clave del proyecto o Numero de CapEx:
TBD
Breve descripción de la solicitud de CapEx:
Equipo de prueba ficticio para validar el extractor
Marca:
TBD
"""


def _construir_hoja_sintetica():
    """Crea un ws con la estructura real (encabezados fila 2, datos fila 3, O/P)."""
    wb = Workbook()
    ws = wb.active
    ws.title = config.HOJA_SEGUIMIENTO  # respeta el nombre con espacio final

    encabezados = {1: "RFQ", 2: "DESCRIPCION", 3: "FECHA DE ARRANQUE",
                   4: "SOLICITANTE", 5: "FECHA LIMITE", 6: "PLANTA",
                   13: "M", 14: "N", 15: "O", 16: "P"}
    for col, txt in encabezados.items():
        ws.cell(row=2, column=col, value=txt)

    # Una fila de datos existente (fila 3), con formulas en O y P como el real.
    ws.cell(row=3, column=1, value="228").number_format = "@"
    ws.cell(row=3, column=2, value="Rectificadora plana")
    ws.cell(row=3, column=3, value=date(2026, 6, 19))
    ws.cell(row=3, column=4, value="Jose Manuel Martinez")
    ws.cell(row=3, column=6, value="Questum Maquinados Ramos")
    ws.cell(row=3, column=13, value=date(2026, 6, 1))   # M
    ws.cell(row=3, column=14, value=date(2026, 6, 25))  # N
    ws.cell(row=3, column=15, value='=IF(N3<=M3,"SI","NO")')  # O
    ws.cell(row=3, column=16, value="=N3-C3")                 # P
    return wb, ws


def _check(cond, msg):
    if not cond:
        raise AssertionError(msg)
    print(f"  OK  {msg}")


def test_extractor_correo():
    print("[1] Extractor de correo Form Approvals")
    _check(parsear_texto("PEDIDO #999 x").get("pedido") == "999", "PEDIDO # -> 999")
    _check(parsear_texto("Request #999 Complete").get("pedido") == "999", "Request # -> 999")

    with tempfile.TemporaryDirectory() as d:
        f = Path(d) / "ficticio_999.txt"
        f.write_text(CORREO_FICTICIO, encoding="utf-8")
        dato = FormApprovalsExtractor().extraer(f)[0]
    _check(dato.rfq == "999", "RFQ = numero de pedido (999)")
    _check(dato.solicitante == "Persona Ficticia Uno", "solicitante desde el formulario, no remitente")
    _check(dato.descripcion.startswith("Equipo de prueba"), "descripcion desde 'Breve descripcion'")
    _check(dato.planta == "Planta Demo Norte", "planta desde 'Unidad de Negocio'")
    _check(dato.fecha_arranque is None, "fecha de arranque queda faltante")
    _check("fecha_arranque" in dato.faltantes, "fecha de arranque reportada como faltante")


def test_actualizar_no_duplica():
    print("[2] Actualizar RFQ existente (no duplica, no toca O/P)")
    wb, ws = _construir_hoja_sintetica()
    dato = RFQData(rfq="228", descripcion="Rectificadora REV-B",
                   solicitante="Jose Manuel Martinez", planta="Questum Maquinados Ramos")
    resultado = maestro.agregar_o_actualizar(ws, dato)
    _check(resultado == "actualizado", "detecta que 228 ya existe -> actualizado")
    _check(ws["A3"].value == "228", "no se duplico (sigue en fila 3)")
    _check(ws["A3"].number_format == "@", "RFQ sigue como texto")
    _check(ws["B3"].value == "Rectificadora REV-B", "descripcion actualizada")
    _check(ws["O3"].value == '=IF(N3<=M3,"SI","NO")', "formula O3 intacta")
    _check(ws["P3"].value == "=N3-C3", "formula P3 intacta")
    _check(maestro.ultima_fila_datos(ws) == 3, "sigue habiendo 1 sola fila de datos")


def test_agregar_traduce_formulas():
    print("[3] Agregar RFQ nuevo (copia y ajusta formulas O/P)")
    wb, ws = _construir_hoja_sintetica()
    dato = RFQData(rfq="185/4", descripcion="Nuevo proyecto",
                   solicitante="Otra Persona", planta="Planta Demo")
    resultado = maestro.agregar_o_actualizar(ws, dato)
    _check(resultado == "agregado", "185/4 no existia -> agregado")
    _check(ws["A4"].value == "185/4", "RFQ con formato '131/4' guardado como texto")
    _check(ws["A4"].number_format == "@", "RFQ nuevo como texto")
    _check(ws["O4"].value == '=IF(N4<=M4,"SI","NO")', "formula O ajustada a fila 4")
    _check(ws["P4"].value == "=N4-C4", "formula P ajustada a fila 4")
    _check(maestro.buscar_fila_rfq(ws, "228") == 3, "228 original intacto en fila 3")


def test_no_toma_encabezado():
    print("[4] No confunde el encabezado 'RFQ' (fila 2) con un dato")
    wb, ws = _construir_hoja_sintetica()
    _check(maestro.buscar_fila_rfq(ws, "RFQ") is None, "buscar 'RFQ' no encuentra la fila de encabezado")


if __name__ == "__main__":
    test_extractor_correo()
    test_actualizar_no_duplica()
    test_agregar_traduce_formulas()
    test_no_toma_encabezado()
    print("\nTODAS LAS VALIDACIONES PASARON.")
